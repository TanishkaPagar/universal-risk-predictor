import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def risk_color(label):
    if "Low" in label:
        return "00C851"
    elif "Medium" in label:
        return "FFD700"
    else:
        return "FF4444"

def generate_excel_report(df_raw, probs, labels):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attrition Risk Report"

    # Header style
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style='thin', color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Add title
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "Employee Attrition Risk Report"
    title_cell.font = Font(bold=True, size=14, color="1F3864")
    title_cell.alignment = Alignment(horizontal='center')

    ws.append([])  # blank row

    headers = ['Employee #', 'Department', 'Job Role', 'Risk Score (%)', 'Risk Level', 'Age']
    ws.append(headers)

    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    for i, (prob, label) in enumerate(zip(probs, labels)):
        row_idx = i + 4
        emp_num = df_raw.iloc[i].get('EmployeeNumber', i + 1)
        dept = df_raw.iloc[i].get('Department', 'N/A')
        role = df_raw.iloc[i].get('JobRole', 'N/A')
        age = df_raw.iloc[i].get('Age', 'N/A')
        score = round(prob * 100, 1)

        row = [emp_num, dept, role, score, label.replace('🟢', '').replace('🟡', '').replace('🔴', '').strip(), age]
        ws.append(row)

        color = risk_color(label)
        fill = PatternFill("solid", fgColor=color)
        for col_num, cell in enumerate(ws[row_idx], 1):
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if col_num == 5:
                cell.fill = fill
                cell.font = Font(bold=True)

    # Column widths
    widths = [14, 20, 28, 16, 16, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Risk Level", "Count", "Percentage"])
    low = sum(1 for l in labels if "Low" in l)
    med = sum(1 for l in labels if "Medium" in l)
    high = sum(1 for l in labels if "High" in l)
    total = len(labels)
    ws2.append(["Low Risk", low, f"{low/total*100:.1f}%"])
    ws2.append(["Medium Risk", med, f"{med/total*100:.1f}%"])
    ws2.append(["High Risk", high, f"{high/total*100:.1f}%"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf